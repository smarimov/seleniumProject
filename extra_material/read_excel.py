import openpyxl

path = "C:\\Users\\shoki\\OneDrive\\Desktop\\Results.xlsx"

excel_file = openpyxl.load_workbook(path)

sheet = excel_file.active

# if you want the specific page:
# sheet = excel_file.get_sheet_by_name("Sheet1")

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(row=r, column=c).value, end='  ')
    print('')
